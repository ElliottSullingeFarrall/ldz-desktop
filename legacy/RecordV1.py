from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox

from openpyxl import *
from os.path import exists

import datetime

hours = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23']
mints = ['00', '05', '10', '15', '20', '25', '30', '35', '40', '45', '50', '55']

departments = ['Prefer not to say', 'Biosciences and Medicine', 'Business School', 'Centre for Environment and Sustainability', 'Chemical and Process Engineering', 'Chemistry', 'Civil and Environmental Engineering', 'Computer Science', 'Economics', 'Electrical and Electronic Engineering', 'Guildford School of Acting', 'Health Sciences', 'Higher Education', 'Hospitality and Tourism Management', 'Law', 'Literature and Languages', 'Mathematics', 'Mechanical Engineering Sciences', 'Music and Media', 'Physics', 'Politics', 'Psychology', 'Sociology', 'Technology Enhanced Learning', 'Veterinary Medicine']
levels      = ['Prefer not to say', 'Foundation', 'UG1', 'UG2', 'UG3', 'UG4', 'Masters', 'PhD', 'Staff']
queries     = ['Maths - Algebra', 'Maths - Calculus', 'Maths - Complex Numbers', 'Maths - Numeracy', 'Maths - Trigonometry', 'Maths - Vector Calculus', 'Maths - Other', 'Software - Excel', 'Software - LaTeX', 'Software - Matlab', 'Software - R', 'Software - SPSS', 'Software - Other', 'Statistics - Data Collection', 'Statistics - Data Presentation', 'Statistics - Statistical Testing', 'Statistics - Statistical Theory', 'Statistics - Other']
types       = ['Drop-in', 'Face-to-face appointment', 'Zoom appointment', 'Significant email thread', 'Drop-in Zoom']
locations   = ['Stag Hill', 'Manor Park', 'Off-campus']

window = Tk()
window.resizable(False, False)
window.title('MASA')

start_hour = StringVar()
start_mint = StringVar()
start_time_label = Label(text='Time In:').grid(row=0, column=0, sticky='e')
start_hour_entry = Spinbox(values=hours, textvariable=start_hour, wrap=True, state='readonly', width=2).grid(row=0, column=1, sticky='w')
start_mint_entry = Spinbox(values=mints, textvariable=start_mint, wrap=True, state='readonly', width=2).grid(row=0, column=2, columnspan=2, sticky='w')

end_hour = StringVar()
end_mint = StringVar()
end_time_label = Label(text='Time Out:').grid(row=1, column=0, sticky='e')
end_hour_entry = Spinbox(values=hours, textvariable=end_hour, wrap=True, state='readonly', width=2).grid(row=1, column=1, sticky='w')
end_mint_entry = Spinbox(values=mints, textvariable=end_mint, wrap=True, state='readonly', width=2).grid(row=1, column=2, columnspan=2, sticky='w')

department = StringVar()
department_label = Label(text='Department:').grid(row=2, column=0, sticky='e')
department_entry = Combobox(textvariable=department, values=departments, state='readonly').grid(row=2, column=1, columnspan=3, sticky='ew')

level = StringVar()
level_label = Label(text='Level:').grid(row=3, column=0, sticky='e')
level_entry = Combobox(textvariable=level, values=levels, state='readonly').grid(row=3, column=1, columnspan=3, sticky='ew')

query = StringVar()
query_label = Label(text='Type of Query:').grid(row=4, column=0, sticky='e')
query_entry = Combobox(textvariable=query, values=queries, state='readonly').grid(row=4, column=1, columnspan=3, sticky='ew')

type = StringVar()
type_label = Label(text='Interaction Type:').grid(row=5, column=0, sticky='e')
type_entry = Combobox(textvariable=type, values=types, state='readonly').grid(row=5, column=1, columnspan=3, sticky='ew')

location = StringVar()
location_label = Label(text='Location:').grid(row=6, column=0, sticky='e')
location_entry = Combobox(textvariable=location, values=locations, state='readonly').grid(row=6, column=1, columnspan=3, sticky='ew')

students = IntVar()
students_label = Label(text='No. of Students').grid(row=7, column=0, sticky='e')
students_entry = Spinbox(textvariable=students, from_=1, to=9, state='readonly', width=1).grid(row=7, column=1, sticky='w')

full_room = StringVar()
full_room_label = Label(text='Was the room full?').grid(row=7, column=2, sticky='e')
full_room_check = Checkbutton(variable=full_room, offvalue='No', onvalue='Yes').grid(row=7, column=3, sticky='e')

def submit_func():
    date = datetime.datetime.now()
    start_time = datetime.datetime.strptime(start_hour.get() +':' + start_mint.get(), '%H:%M')
    end_time = datetime.datetime.strptime(end_hour.get() +':' + end_mint.get(), '%H:%M')
    diff_time = end_time - start_time

    if exists('data.xlsx'):
        workbook = load_workbook(filename='data.xlsx')
        sheet = workbook.active
        sheet.append([date.strftime('%d/%m/%Y'), start_time.strftime('%H:%M'), end_time.strftime('%H:%M'), str(diff_time)[:-3], department.get(), level.get(), query.get(), type.get(), location.get(), students.get(), full_room.get()])
        workbook.save(filename='data.xlsx')
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Date', 'Time In', 'Time Out', 'Time of Session', 'Department', 'Level', 'Type of Query', 'Interaction Type', 'Campus', 'Number of identical queries', 'Room Full?'])
        sheet.append([date.strftime('%d/%m/%Y'), start_time.strftime('%H:%M'), end_time.strftime('%H:%M'), str(diff_time)[:-3], department.get(), level.get(), query.get(), type.get(), location.get(), students.get(), full_room.get()])
        workbook.save(filename='data.xlsx')

    messages['text'] = 'Appointment submitted from ' + start_hour.get() +':' + start_mint.get() + ' to ' + end_hour.get() + ':' + end_mint.get()
    undo_button['state'] = NORMAL
    init_vars()
        
submit_button = Button(text='Submit', command=submit_func)
submit_button.grid(row=8, column=0, columnspan=2)

def undo_func():
    confirmation = messagebox.askquestion('', 'Are you sure you would like to undo?')
    if confirmation == 'no':
        return

    workbook = load_workbook(filename='data.xlsx')
    sheet = workbook.active
    if sheet.max_row > 1:
        start_time_cell = sheet.cell(row=sheet.max_row, column=2)
        end_time_cell = sheet.cell(row=sheet.max_row, column=3)
        sheet.delete_rows(sheet.max_row, 1)
        messages['text'] = 'Appointment removed from ' + str(start_time_cell.value) + ' to ' + str(end_time_cell.value)
    else:
        undo_button['state'] = DISABLED
        messages['text'] = 'Undo failed: No entries left to remove.'
    workbook.save(filename='data.xlsx')

undo_button = Button(text='Undo', command=undo_func, state=DISABLED)
undo_button.grid(row=8, column=2, columnspan=2)

messages = Label(text='Welcome to the MASA appointment manager.')
messages.grid(row=9, column=0, columnspan=4)

def init_vars():
    start_hour.set('12')
    start_mint.set('00')
    end_hour.set('12')
    end_mint.set('00')
    department.set('')
    level.set('')
    query.set('')
    type.set(types[0])
    location.set(locations[0])
    students.set(1)
    full_room.set('No')

init_vars()
window.mainloop()