from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry

from openpyxl import *
from os.path import exists

import datetime

hours = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23']
mints = ['00', '05', '10', '15', '20', '25', '30', '35', '40', '45', '50', '55']

#departments = ['Prefer not to say', 'Biosciences and Medicine', 'Business School', 'Centre for Environment and Sustainability', 'Chemical and Process Engineering', 'Chemistry', 'Civil and Environmental Engineering', 'Computer Science', 'Economics', 'Electrical and Electronic Engineering', 'Guildford School of Acting', 'Health Sciences', 'Higher Education', 'Hospitality and Tourism Management', 'Law', 'Literature and Languages', 'Mathematics', 'Mechanical Engineering Sciences', 'Music and Media', 'Physics', 'Politics', 'Psychology', 'Sociology', 'Technology Enhanced Learning', 'Veterinary Medicine']
#levels      = ['Prefer not to say', 'Foundation', 'UG1', 'UG2', 'UG3', 'UG4', 'Masters', 'PhD', 'Staff']
#queries     = ['Maths - Algebra', 'Maths - Calculus', 'Maths - Complex Numbers', 'Maths - Numeracy', 'Maths - Trigonometry', 'Maths - Vector Calculus', 'Maths - Other', 'Software - Excel', 'Software - LaTeX', 'Software - Matlab', 'Software - R', 'Software - SPSS', 'Software - Other', 'Statistics - Data Collection', 'Statistics - Data Presentation', 'Statistics - Statistical Testing', 'Statistics - Statistical Theory', 'Statistics - Other']
#types       = ['Drop-in', 'Face-to-face appointment', 'Zoom appointment', 'Significant email thread', 'Drop-in Zoom']
#locations   = ['Stag Hill', 'Manor Park', 'Off-campus']

config_file = load_workbook(filename='config.xlsx')
config = config_file.active
departments = [cell.value for cell in config['A']]
levels      = [cell.value for cell in config['B']]
queries     = [cell.value for cell in config['C']]
types       = [cell.value for cell in config['D']]
locations   = [cell.value for cell in config['E']]
config_file.save(filename='config.xlsx')

window = Tk()
window.resizable(False, False)
window.title('MASA')

date = StringVar()
date_label = ttk.Label(text='Date:').grid(row=0, column=0, sticky='e')
date_entry = DateEntry(date_pattern='dd/mm/yyyy', selectmode = 'day', textvariable=date, state='readonly').grid(row=0, column=1, columnspan=3, sticky='we')

start_hour = StringVar()
start_mint = StringVar()
start_time_label = ttk.Label(text='Time In:').grid(row=1, column=0, sticky='e')
start_hour_entry = ttk.Spinbox(values=hours, textvariable=start_hour, wrap=True, state='readonly', width=3).grid(row=1, column=1, sticky='w')
start_mint_entry = ttk.Spinbox(values=mints, textvariable=start_mint, wrap=True, state='readonly', width=3).grid(row=1, column=2, columnspan=2, sticky='w')

end_hour = StringVar()
end_mint = StringVar()
end_time_label = ttk.Label(text='Time Out:').grid(row=2, column=0, sticky='e')
end_hour_entry = ttk.Spinbox(values=hours, textvariable=end_hour, wrap=True, state='readonly', width=3).grid(row=2, column=1, sticky='w')
end_mint_entry = ttk.Spinbox(values=mints, textvariable=end_mint, wrap=True, state='readonly', width=3).grid(row=2, column=2, columnspan=2, sticky='w')

department = StringVar()
department_label = ttk.Label(text='Department:').grid(row=3, column=0, sticky='e')
department_entry = ttk.Combobox(textvariable=department, values=departments, state='readonly').grid(row=3, column=1, columnspan=3, sticky='ew')

level = StringVar()
level_label = ttk.Label(text='Level:').grid(row=4, column=0, sticky='e')
level_entry = ttk.Combobox(textvariable=level, values=levels, state='readonly').grid(row=4, column=1, columnspan=3, sticky='ew')

query = StringVar()
query_label = ttk.Label(text='Type of Query:').grid(row=5, column=0, sticky='e')
query_entry = ttk.Combobox(textvariable=query, values=queries, state='readonly').grid(row=5, column=1, columnspan=3, sticky='ew')

type = StringVar()
type_label = ttk.Label(text='Interaction Type:').grid(row=6, column=0, sticky='e')
type_entry = ttk.Combobox(textvariable=type, values=types, state='readonly').grid(row=6, column=1, columnspan=3, sticky='ew')

location = StringVar()
location_label = ttk.Label(text='Location:').grid(row=7, column=0, sticky='e')
location_entry = ttk.Combobox(textvariable=location, values=locations, state='readonly').grid(row=7, column=1, columnspan=3, sticky='ew')

students = IntVar()
students_label = ttk.Label(text='No. of Students').grid(row=8, column=0, sticky='e')
students_entry = ttk.Spinbox(textvariable=students, from_=1, to=9, state='readonly', width=2).grid(row=8, column=1, sticky='w')

full_room = StringVar()
full_room_label = ttk.Label(text='Was the room full?').grid(row=8, column=2, sticky='e')
full_room_check = ttk.Checkbutton(variable=full_room, offvalue='No', onvalue='Yes').grid(row=8, column=3, sticky='e')

def submit_func():
    if '' in get_vars().values():
        messagebox.showinfo(message='Please fill in all the fields')
        return
    if exists('data.xlsx'):
        data_file = load_workbook(filename='data.xlsx')
        data = data_file.active
        data.append(list(get_vars().values()))
        data_file.save(filename='data.xlsx')
    else:
        data_file = Workbook()
        data_file.security.lockStructure = True
        data = data_file.active
        data.protection.enable()
        data.append(list(get_vars().keys()))
        data.append(list(get_vars().values()))
        data_file.save(filename='data.xlsx')
    init_vars()        
submit_button = ttk.Button(text='Submit', command=submit_func)
submit_button.grid(row=9, column=0, columnspan=2)

def del_func():
    del_window = Toplevel(window)
    del_window.resizable(False, False)
    del_window.title('Data')
    del_window.attributes('-topmost', 'true')
    del_window.grab_set()

    data_file = load_workbook(filename='data.xlsx')
    data = data_file.active
    table = ttk.Treeview(del_window, columns=[cell.value for cell in data[1]], show='headings', selectmode='browse')
    data_file.save(filename='data.xlsx')

    for col in table["column"]:
        table.heading(col, text=col, anchor=CENTER)
        table.column(col, stretch=False, width=100, anchor=CENTER)
    for row in data.iter_rows(min_row=2):
        table.insert("", "end", values=[cell.value for cell in row])
    table.grid(row=0, column=0)

    vsb = ttk.Scrollbar(del_window, orient="vertical", command=table.yview)
    vsb.grid(row=0, column=1, sticky='ns')
    table.configure(yscrollcommand=vsb.set)

    def del_selection(event):
        confirmation = messagebox.askquestion(parent=del_window, message='Are you sure you would like to delete this entry?')
        if confirmation == 'no':
            return
        row_num = table.index(table.selection()) + 2
        data_file = load_workbook(filename='data.xlsx')
        data = data_file.active
        data.delete_rows(row_num, 1)
        data_file.save(filename='data.xlsx')
        del_window.destroy()
    table.bind("<<TreeviewSelect>>", del_selection)
del_button = ttk.Button(text='Delete...', command=del_func)
del_button.grid(row=9, column=2, columnspan=2)

def init_vars():
    date.set(datetime.datetime.now().strftime("%d/%m/%Y"))
    start_hour.set('12')
    start_mint.set('00')
    end_hour.set('12')
    end_mint.set('00')
    department.set('')
    level.set('')
    query.set('')
    type.set(types[0])
    location.set(locations[0])
    students.set(1)
    full_room.set('No')

def get_vars():
    start_time = datetime.datetime.strptime(start_hour.get() +':' + start_mint.get(), '%H:%M')
    end_time = datetime.datetime.strptime(end_hour.get() +':' + end_mint.get(), '%H:%M')
    diff_time = end_time - start_time
    return {'Date' : date.get(), 'Time In' : start_time.strftime('%H:%M'), 'Time Out' : end_time.strftime('%H:%M'), 'Time of Session' : str(diff_time)[:-3], 'Department' : department.get(), 'Level' : level.get(), 'Type of Query' : query.get(), 'Interaction Type' : type.get(), 'Campus' : location.get(), 'Number of identical queries' : students.get(), 'Room Full?' : full_room.get()}

def load_config(filename):
    if exists(filename):
        try:
            config_file = load_workbook(filename=filename)
        except PermissionError:
            answer = messagebox.askretrycancel('', 'Unable to load the config file. Please try again.')
            if answer:
                load_config(filename)
            else:
                window.destroy()
    else:
        answer = messagebox.showerror(message='No config file found.')
        if answer:
            window.destroy()
    return config_file
    
def load_data(filename):
    if exists(filename):
        try:
            data_file = load_workbook(filename=filename)
        except PermissionError:
            answer = messagebox.askretrycancel('', 'Unable to load the data file. Please try again.')
            if answer:
                load_data(filename)
            else:
                window.destroy()
    else:
        data_file = Workbook()
        data_file.security.lockStructure = True
        data = data_file.active
        data.protection.enable()
        data.append(list(get_vars().keys()))
    return data_file

init_vars()
window.mainloop()