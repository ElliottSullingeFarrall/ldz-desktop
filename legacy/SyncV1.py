import tkinter as tk
from tkinter import ttk

import openpyxl as xl
import os

window = tk.Tk()
window.resizable(False, False)
window.title('Sync')
window.iconphoto(True, tk.PhotoImage(file='images/stag.png'))

os.chdir('Input')
files = os.listdir()

files_list = tk.Listbox(selectmode=tk.MULTIPLE)
for file in files:
    files_list.insert(tk.END, file)
files_list.grid(row=0, column=0, columnspan=2, sticky='ew')

def select_all():
    files_list.select_set(0, tk.END)
select_button = ttk.Button(text='Select All', command=select_all)
select_button.grid(row=1, column=0, sticky='ew')

def sync():
    sync_button['state']   = tk.DISABLED
    select_button['state'] = tk.DISABLED
    files_list.bindtags((files_list, window, "all"))

    output = xl.Workbook()

    def get_index(idx):
        if idx == 0:
            return 1
        else:
            return 2

    files = [files_list.get(idx) for idx in files_list.curselection()]
    for idx, filename in enumerate(files):
        data = xl.load_workbook(filename=filename)
        for row in data.active.iter_rows(min_row=get_index(idx)):
            row_values = [cell.value for cell in row]
            output.active.append(row_values)
        data.save(filename=filename)
        files_list.selection_clear(idx)
        window.update()

    os.chdir('..')
    output.save(filename='Output.xlsx')
    window.destroy()
sync_button = ttk.Button(text='Sync', command=sync)
sync_button.grid(row=1, column=1, sticky='ew')

window.mainloop()