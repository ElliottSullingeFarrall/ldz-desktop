import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry

from datetime import datetime

import openpyxl as xl
import os
import sys

class Config:
    def __init__(self, filename):
        self.filename = filename

    def __enter__(self):
        if os.path.exists(self.filename):
            try:
                self.wb = xl.load_workbook(filename=self.filename)
                self.ws = self.wb.active
            except PermissionError:
                answer = messagebox.askretrycancel(message='Unable to load the config file. Please try again.')
                if answer:
                    self.__enter__()
                else:
                    window.destroy()
        else:
            answer = messagebox.showerror(message='No config file found.')
            if answer:
                window.destroy()
        return self.ws

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            self.wb.save(filename=self.filename)
            self.wb.close()
        except PermissionError:
            answer = messagebox.askretrycancel(message='Unable to save the config file. Please try again.')
            if answer:
                self.__exit__(exc_type, exc_val, exc_tb)
            else:
                window.destroy()

class Data:
    def __init__(self, filename):
        self.filename = filename

    def __enter__(self):
        if os.path.exists(self.filename):
            try:
                self.wb = xl.load_workbook(filename=self.filename)
                self.ws = self.wb.active
            except PermissionError:
                answer = messagebox.askretrycancel(message='Unable to load the data file. Please try again.')
                if answer:
                    self.__enter__()
                else:
                    window.destroy()
        else:
            self.wb = xl.Workbook()
            self.wb.security.lockStructure = True
            self.wb.active.protection.enable()
            self.wb.active.append(list(fields.get().keys()))
            self.ws = self.wb.active
        return self.ws

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            self.wb.save(filename=self.filename)
            self.wb.close()
        except PermissionError:
            answer = messagebox.askretrycancel(message='Unable to save the data file. Please try again.')
            if answer:
                self.__exit__(exc_type, exc_val, exc_tb)
            else:
                window.destroy()

class Fields:
    def __init__(self, config):
        self.labels     =  ['Date', 'Time In', 'Time Out', 'Time of Session'] + [col[0] for col in config.iter_cols(values_only=True)] + ['Number of Idenitical Queries', 'Room Full?']
        self.date       =  {                 'var' : tk.StringVar()}
        self.start_hour =  {                 'var' : tk.StringVar(), 'values' : [f"{hour:02}" for hour in range(0, 24, 1)], 'default' : '12'  }
        self.start_mint =  {                 'var' : tk.StringVar(), 'values' : [f"{mint:02}" for mint in range(0, 60, 5)], 'default' : '00'  }
        self.end_hour   =  {                 'var' : tk.StringVar(), 'values' : [f"{hour:02}" for hour in range(0, 24, 1)], 'default' : '12'  }
        self.end_mint   =  {                 'var' : tk.StringVar(), 'values' : [f"{mint:02}" for mint in range(0, 60, 5)], 'default' : '00'  }
        self.add_fields = [{'name' : col[0], 'var' : tk.StringVar(), 'values' : list(filter(None, col[1:])),                'default' : col[1]} for col in config.iter_cols(values_only=True)]
        self.students   =  {                 'var' : tk.IntVar(),    'values' : tuple(range(1,10)),                         'default' : 1     }
        self.full_room  =  {                 'var' : tk.StringVar(),                                                        'default' : 'No'  }
        self.refresh()

    def refresh(self):
        self.date['var'].set(f"{datetime.now():%d/%m/%Y}")
        self.start_hour['var'].set(self.start_hour['default'])
        self.start_mint['var'].set(self.start_mint['default'])
        self.end_hour['var'].set(self.end_hour['default'])
        self.end_mint['var'].set(self.end_mint['default'])
        for field in self.add_fields:
            field['var'].set(field['default'] or '')
        self.students['var'].set(self.students['default'])
        self.full_room['var'].set(self.full_room['default'])

    def get(self):
        start_time = datetime.strptime(f"{self.start_hour['var'].get()}:{self.start_mint['var'].get()}", '%H:%M')
        end_time = datetime.strptime(f"{self.end_hour['var'].get()}:{self.end_mint['var'].get()}", '%H:%M')
        diff_time = end_time - start_time
        return [self.date['var'].get(), f"{start_time:%H:%M}", f"{end_time:%H:%M}", str(diff_time)[:-3]] + [field['var'].get() for field in self.add_fields] + [self.students['var'].get(), self.full_room['var'].get()]

def main():
    path = os.path.dirname(sys.argv[0])
    os.chdir(path)
    while 'Record.app' in path:
        path = os.path.dirname(path)
        os.chdir(path)
    try:
        image = f'{sys._MEIPASS}/images/stag.png'
    except AttributeError:
        image = f'images/stag.png'

    window = tk.Tk()
    window.title('Record')
    window.iconphoto(True, tk.PhotoImage(file=image))
    window.resizable(False, False)

    with Config('config.xlsx') as config:
        fields = Fields(config)

    ttk.Label(text='Date:').grid(row=0, column=0, sticky='e')
    DateEntry(textvariable=fields.date['var'], date_pattern='dd/mm/yyyy', selectmode = 'day', state='readonly').grid(row=0, column=1, columnspan=3, sticky='we')

    ttk.Label(text='Time In:').grid(row=1, column=0, sticky='e')
    ttk.Spinbox(textvariable=fields.start_hour['var'], values=fields.start_hour['values'], state='readonly', wrap=True, width=3).grid(row=1, column=1, sticky='w')
    ttk.Spinbox(textvariable=fields.start_mint['var'], values=fields.start_mint['values'], state='readonly', wrap=True, width=3).grid(row=1, column=2, columnspan=2, sticky='w')

    ttk.Label(text='Time Out:').grid(row=2, column=0, sticky='e')
    ttk.Spinbox(textvariable=fields.end_hour['var'], values=fields.end_hour['values'], state='readonly', wrap=True, width=3).grid(row=2, column=1, sticky='w')
    ttk.Spinbox(textvariable=fields.end_mint['var'], values=fields.end_hour['values'], state='readonly', wrap=True, width=3).grid(row=2, column=2, columnspan=2, sticky='w')

    for pos, field in enumerate(fields.add_fields):
        ttk.Label(text=f"{field['name']}:").grid(row=3+pos, column=0, sticky='e')
        ttk.Combobox(textvariable=field['var'], values=field['values'], state='readonly').grid(row=3+pos, column=1, columnspan=3, sticky='ew')

    ttk.Label(text='No. of Students:').grid(row=4+pos, column=0, sticky='e')
    ttk.Spinbox(textvariable=fields.students['var'], values=fields.students['values'], state='readonly', width=2).grid(row=4+pos, column=1, sticky='w')

    ttk.Label(text='Was the room full?').grid(row=4+pos, column=2, sticky='e')
    ttk.Checkbutton(variable=fields.full_room['var'], offvalue='No', onvalue='Yes').grid(row=4+pos, column=3, sticky='e')

    def submit_func():
        if '' in fields.get():
            messagebox.showinfo(message='Please fill in all the fields.')
        else:
            with Data('data.xlsx') as data:
                data.append(fields.get())
            fields.refresh()       
    submit_button = ttk.Button(text='Submit', command=submit_func)
    submit_button.grid(row=5+pos, column=0, columnspan=2)

    def del_func():
        del_window = tk.Toplevel(window)
        del_window.title('Data')
        del_window.resizable(False, False)
        del_window.attributes('-topmost', 'true')
        del_window.grab_set()

        with Data('data.xlsx') as data:
            table = ttk.Treeview(del_window, columns=fields.labels, show='headings', selectmode='browse')
            for col in table['columns']:
                table.heading(col, text=col, anchor=tk.CENTER)
                table.column(col, stretch=False, anchor=tk.CENTER, width=100)
            for row in data.iter_rows(min_row=2, values_only=True):
                table.insert(parent='', index='end', values=row)
        table.grid(row=0, column=0)

        vsb = ttk.Scrollbar(del_window, orient="vertical", command=table.yview)
        vsb.grid(row=0, column=1, sticky='ns')
        table.configure(yscrollcommand=vsb.set)

        def del_selection(event):
            confirmation = messagebox.askquestion(parent=del_window, message='Are you sure you would like to delete this entry?')
            if confirmation == 'no':
                pass
            else:
                row_num = table.index(table.selection()) + 2
                with Data('data.xlsx') as data:
                    data.delete_rows(row_num, 1)
                del_window.destroy()
        table.bind("<<TreeviewSelect>>", del_selection)
    del_button = ttk.Button(text='Delete...', command=del_func)
    del_button.grid(row=5+pos, column=2, columnspan=2)

    window.mainloop()

if __name__ == '__main__':
    main()