import openpyxl as xl
import sys

class Output:
    def __init__(self, filename):
        self.filename = filename

    def __enter__(self):
        try:
            self.wb = xl.Workbook()
            self.table = self.wb.active
        except PermissionError:
            raise Exception("Unable to load the output file. Please try again.")
        return self.table

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            self.wb.save(filename=self.filename)
            self.wb.close()
        except PermissionError:
            raise Exception("Unable to save the output file. Please try again.")

class Data:
    def __init__(self, filename):
        self.filename = filename

    def __enter__(self):
        try:
            self.wb = xl.Workbook()
            self.table = self.wb.active
        except PermissionError:
            raise Exception(f"Unable to load the data file ({self.filename}). Please try again.")
        return self.table

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            self.wb.save(filename=self.filename)
            self.wb.close()
        except PermissionError:
            raise Exception(f"Unable to save the data file ({self.filename}). Please try again.")

def main():
    files = sys.argv[1:]
    files = ['test_data/1.xlsx', 'test_data/2.xlsx', 'test_data/3.xlsx']

    with Output("output.xlsx") as output:
        for idx, filename in enumerate(files):
            with Data(filename) as data:
                for row in data.iter_rows(min_row=1 if idx == 0 else 2, values_only=True):
                    output.append(row)
    exit
if __name__ == '__main__':
    main()